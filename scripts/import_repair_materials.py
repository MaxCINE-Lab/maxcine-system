#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MATERIAL_START_ROW = 5
MATERIAL_END_ROW = 50
POLICY_START_ROW = 59
POLICY_END_ROW = 70


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def cents(value: Any) -> int | None:
    if value is None or text(value) == "":
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value) * 100))
    raw = text(value)
    if raw == "不适用":
        return None
    try:
        return int(round(float(raw) * 100))
    except ValueError:
        return None


def sql(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def stable_id(prefix: str, value: str) -> str:
    clean = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    digest = hashlib.sha1(value.encode("utf-8")).hexdigest()[:8]
    return f"{prefix}-{clean or 'row'}-{digest}"


def split_codes(raw: str) -> list[str]:
    raw = raw.replace("，", "/").replace(",", "/")
    result: list[str] = []
    for part in raw.split("/"):
        part = part.strip()
        if not part:
            continue
        if re.match(r"^W\d{3}$", part, re.I):
            part = f"CG.{part.upper()}"
        if re.match(r"^CG\.\d{4}$", part, re.I):
            result.append(part.upper())
        elif re.match(r"^CG\.W\d{3}$", part, re.I):
            result.append(part.upper())
        elif re.match(r"^CG\.W\d{3}-CG\.W\d{3}$", part, re.I):
            result.append(part.upper())
        else:
            result.append(part)
    return result


def parse_warranty(raw: str) -> tuple[int | None, dict[str, Any]]:
    if not raw:
        return None, {}
    if raw == "90天":
        return 90, {"type": "fixed_days", "days": 90, "raw": raw}
    if raw == "180天":
        return 180, {"type": "fixed_days", "days": 180, "raw": raw}
    if raw == "1年":
        return 365, {"type": "fixed_days", "days": 365, "raw": raw}
    if raw == "30天":
        return 30, {"type": "fixed_days", "days": 30, "raw": raw}
    if "付费30天" in raw and "无保修" in raw:
        return None, {"type": "conditional", "included": "none", "paid_days": 30, "raw": raw}
    return None, {"type": "manual_confirm", "raw": raw}


def parse_price(value: Any) -> tuple[int | None, str]:
    if value is None or text(value) == "":
        return None, "missing"
    if text(value) == "不适用":
        return None, "not_applicable"
    value_cents = cents(value)
    if value_cents is None:
        return None, "manual_confirm"
    return value_cents, "zero" if value_cents == 0 else "available"


def parse_service_fee(value: Any) -> tuple[int | None, str, dict[str, Any]]:
    raw = text(value)
    if raw == "":
        return None, "missing", {}
    if raw == "不适用":
        return None, "not_applicable", {"raw": raw}
    if raw == "物料价格含服务费":
        return None, "included", {"raw": raw, "included_in_material_price": True}
    match = re.search(r"标准\s*(\d+)\s*增强\s*(\d+)", raw)
    if match:
        return None, "version_rule", {
            "raw": raw,
            "standard": int(match.group(1)) * 100,
            "enhanced": int(match.group(2)) * 100,
        }
    value_cents = cents(value)
    if value_cents is None:
        return None, "text_rule", {"raw": raw}
    return value_cents, "zero" if value_cents == 0 else "fixed", {}


def policy_map(ws) -> dict[str, list[dict[str, Any]]]:
    policies: dict[str, list[dict[str, Any]]] = {}
    for row in range(POLICY_START_ROW, POLICY_END_ROW + 1):
        material_name = text(ws.cell(row, 2).value)
        raw_codes = text(ws.cell(row, 3).value)
        warranty_raw = text(ws.cell(row, 4).value)
        if not material_name or not raw_codes or not warranty_raw:
            continue
        days, rule = parse_warranty(warranty_raw)
        for code in split_codes(raw_codes):
            policies.setdefault(code.upper(), []).append({
                "name": material_name,
                "days": days,
                "rule": rule,
                "source_row": row,
            })
    return policies


def recommendation_triggers(name: str) -> list[tuple[str, str, int]]:
    rules = [
        ("外框", "fault_part", "镜头主上组件外框"),
        ("主上模块", "fault_part", "镜头主上模块"),
        ("围脖", "fault_part", "环绕围脖"),
        ("主体框架", "fault_part", "主体框架"),
        ("光学镜片", "fault_part", "主体光学镜片"),
        ("长焦外光学镜片", "fault_part", "长焦外光学镜片"),
        ("配重", "fault_part", "配重模块"),
        ("ND8", "fault_part", "ND8组件"),
        ("ND16", "fault_part", "ND16组件"),
        ("ND32", "fault_part", "ND32组件"),
        ("ND64", "fault_part", "ND64组件"),
        ("UV", "fault_part", "外置UV"),
        ("收纳盒", "fault_part", "收纳盒"),
        ("背胶", "damage_type", "粘合失效"),
        ("卡扣", "damage_type", "卡扣损坏"),
        ("镜片", "damage_type", "镜片划伤"),
    ]
    return [(trigger_type, trigger_value, index + 1) for index, (needle, trigger_type, trigger_value) in enumerate(rules) if needle in name]


def parse_materials(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    policies = policy_map(ws)
    materials: list[dict[str, Any]] = []
    stats = {"valid_rows": 0, "with_code": 0, "missing_code": 0, "warning": 0, "error": 0}
    seen_codes: set[str] = set()
    duplicate_codes: set[str] = set()

    for row in range(MATERIAL_START_ROW, MATERIAL_END_ROW + 1):
        sequence = text(ws.cell(row, 1).value)
        material_name = text(ws.cell(row, 2).value)
        if not sequence and not material_name:
            continue
        code = text(ws.cell(row, 3).value).upper()
        applicable_models = text(ws.cell(row, 4).value)
        note = text(ws.cell(row, 5).value)
        price, price_status = parse_price(ws.cell(row, 6).value)
        service_fee, service_status, service_rule = parse_service_fee(ws.cell(row, 7).value)
        retail_category = text(ws.cell(row, 8).value)
        can_replace = text(ws.cell(row, 9).value) in {"是", "Y", "YES", "1", "可"}
        issues: list[dict[str, str]] = []

        if not code:
            issues.append({"severity": "error", "code": "missing_material_code", "message": "缺少料号，已保留为异常物料，需人工处理。"})
        elif code in seen_codes:
            duplicate_codes.add(code)
            issues.append({"severity": "error", "code": "duplicate_material_code", "message": "料号重复，需人工核对。"})
        else:
            seen_codes.add(code)

        if price_status == "not_applicable":
            issues.append({"severity": "warning", "code": "price_not_applicable", "message": "保外价格为“不适用”，不得自动报价。"})
        if price_status == "missing":
            issues.append({"severity": "warning", "code": "price_missing", "message": "缺少保外价格，需管理员确认。"})
        if service_status in {"text_rule", "version_rule", "included", "not_applicable", "missing"}:
            issues.append({"severity": "warning", "code": f"service_fee_{service_status}", "message": "服务费需要按备注或管理员规则确认。"})
        if not applicable_models:
            issues.append({"severity": "warning", "code": "applicable_model_missing", "message": "适用型号为空，默认不按型号优先推荐。"})
        if re.search(r"CG\.0\d{3}", note):
            issues.append({"severity": "warning", "code": "note_code_format_mismatch", "message": "备注中的料号格式与正式料号格式不一致，原文已保留。"})

        code_policies = policies.get(code, []) if code else []
        warranty_days = None
        warranty_rule: dict[str, Any] = {}
        warranty_policy = ""
        if code_policies:
            unique_days = {item["days"] for item in code_policies}
            if len(unique_days) == 1:
                warranty_days = next(iter(unique_days))
                warranty_rule = code_policies[0]["rule"]
                warranty_policy = text(code_policies[0]["rule"].get("raw", ""))
            else:
                warranty_rule = {"type": "conditional_by_product", "conditions": code_policies}
                warranty_policy = "按销售产品版本确认"
                issues.append({"severity": "warning", "code": "conditional_warranty_policy", "message": "同一维修物料对应多个保修期限，需按销售产品版本确认。"})

        status = "error" if any(item["severity"] == "error" for item in issues) else "warning" if issues else "normal"
        stats["valid_rows"] += 1
        stats["with_code" if code else "missing_code"] += 1
        stats[status if status in {"warning", "error"} else "valid_rows"] = stats.get(status, 0) + (1 if status in {"warning", "error"} else 0)

        raw = {
            "序号": sequence,
            "物料名称": material_name,
            "料号P/N": code,
            "适用型号": applicable_models,
            "备注": note,
            "保外价格": text(ws.cell(row, 6).value),
            "保外服务费": text(ws.cell(row, 7).value),
            "零售产品": retail_category,
            "可当全套换": text(ws.cell(row, 9).value),
        }
        material_id = stable_id("repair-material", code or f"row-{row}-{material_name}")
        materials.append({
            "id": material_id,
            "row": row,
            "sequence": sequence,
            "material_code": code or None,
            "material_name": material_name,
            "applicable_models": applicable_models,
            "description": note,
            "price": price,
            "price_status": price_status,
            "service_fee": service_fee,
            "service_status": service_status,
            "service_rule": service_rule,
            "retail_category": retail_category,
            "can_replace": can_replace,
            "warranty_policy": warranty_policy,
            "warranty_days": warranty_days,
            "warranty_rule": warranty_rule,
            "source_note": note,
            "raw": raw,
            "issues": issues,
            "quality": status,
        })

    for material in materials:
        if material["material_code"] in duplicate_codes:
            material["quality"] = "error"
    return materials, stats


def build_sql(materials: list[dict[str, Any]], source: Path) -> str:
    fingerprint = hashlib.sha256(source.read_bytes()).hexdigest()
    batch_id = stable_id("repair-material-import", fingerprint)
    warning_rows = sum(1 for item in materials if item["quality"] == "warning")
    error_rows = sum(1 for item in materials if item["quality"] == "error")
    lines = [
        "PRAGMA foreign_keys = ON;",
        "BEGIN TRANSACTION;",
        (
            "INSERT INTO repair_material_import_batches "
            "(id, source_filename, source_file_fingerprint, source_sheet, total_rows, imported_rows, skipped_rows, warning_rows, error_rows) VALUES "
            f"({sql(batch_id)}, {sql(source.name)}, {sql(fingerprint)}, 'Sheet1', {len(materials)}, {len(materials) - error_rows}, {error_rows}, {warning_rows}, {error_rows}) "
            "ON CONFLICT(source_file_fingerprint) DO UPDATE SET "
            "total_rows=excluded.total_rows, imported_rows=excluded.imported_rows, skipped_rows=excluded.skipped_rows, warning_rows=excluded.warning_rows, error_rows=excluded.error_rows;"
        ),
    ]
    for item in materials:
        active = 0 if item["quality"] == "error" else 1
        columns = [
            "id", "material_code", "material_name", "applicable_models", "description",
            "out_of_warranty_price_cents", "price_status", "out_of_warranty_service_fee_cents", "service_fee_status", "service_fee_rule_json",
            "retail_category", "can_replace_as_whole_set", "warranty_policy", "warranty_days", "warranty_rule_json",
            "active", "source", "source_row_number", "source_note", "source_raw_json", "data_quality_status", "issues_json",
        ]
        values = [
            item["id"], item["material_code"], item["material_name"], item["applicable_models"], item["description"],
            item["price"], item["price_status"], item["service_fee"], item["service_status"], json.dumps(item["service_rule"], ensure_ascii=False),
            item["retail_category"], 1 if item["can_replace"] else 0, item["warranty_policy"], item["warranty_days"], json.dumps(item["warranty_rule"], ensure_ascii=False),
            active, "售后物料.xlsx", item["row"], item["source_note"], json.dumps(item["raw"], ensure_ascii=False), item["quality"], json.dumps(item["issues"], ensure_ascii=False),
        ]
        assignments = ", ".join([f"{column}=excluded.{column}" for column in columns[2:]])
        lines.append(
            f"INSERT INTO repair_materials ({', '.join(columns)}) VALUES ({', '.join(sql(value) for value in values)}) "
            f"ON CONFLICT(id) DO UPDATE SET {assignments}, updated_at=CURRENT_TIMESTAMP;"
        )
        for trigger_type, trigger_value, priority in recommendation_triggers(item["material_name"]):
            rec_id = stable_id("repair-rec", f"{item['id']}-{trigger_type}-{trigger_value}")
            lines.append(
                "INSERT OR IGNORE INTO repair_material_recommendations "
                "(id, material_id, trigger_type, trigger_value, priority, note) VALUES "
                f"({sql(rec_id)}, {sql(item['id'])}, {sql(trigger_type)}, {sql(trigger_value)}, {priority}, '售后物料初始化推荐');"
            )
    lines.append("COMMIT;")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import MaxCINE repair materials into local D1.")
    parser.add_argument("--file", default="/Users/rog/Desktop/售后物料.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--persist-to", default="work/gsx-d1")
    args = parser.parse_args()
    source = Path(args.file).expanduser().resolve()
    if not source.exists():
        raise SystemExit(f"文件不存在：{source}")
    materials, stats = parse_materials(source)
    sql_text = build_sql(materials, source)
    print(json.dumps({
        "source": str(source),
        "validMaterialRows": len(materials),
        "withMaterialCode": sum(1 for item in materials if item["material_code"]),
        "missingMaterialCode": sum(1 for item in materials if not item["material_code"]),
        "normalRows": sum(1 for item in materials if item["quality"] == "normal"),
        "warningRows": sum(1 for item in materials if item["quality"] == "warning"),
        "errorRows": sum(1 for item in materials if item["quality"] == "error"),
        "anomalies": [
            {"row": item["row"], "materialCode": item["material_code"], "materialName": item["material_name"], "issues": item["issues"]}
            for item in materials if item["issues"]
        ],
    }, ensure_ascii=False, indent=2))
    if args.dry_run:
        return
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False, encoding="utf-8") as handle:
        handle.write(sql_text)
        temp_sql = handle.name
    try:
        subprocess.run([
            "npx", "wrangler", "d1", "execute", "maxcine-db",
            "--local", "--persist-to", args.persist_to,
            "--file", temp_sql,
            "--config", "apps/api/wrangler.toml",
        ], check=True)
    finally:
        Path(temp_sql).unlink(missing_ok=True)


if __name__ == "__main__":
    main()
